from openpyxl import *
import easygui
import ast

class Horizon(object):
    def __init__(self):
        super().__init__()
        tessname = easygui.fileopenbox(msg="Tess File", title="Horizon Tess/QB Compare", default="*.xlsx")
        self.__tessWB = load_workbook(tessname, read_only=True)
        qbname = easygui.fileopenbox(msg="QB File", title="Horizon Tess/QB Compare", default="*.xlsx")
        self.__qbWB = load_workbook(qbname, read_only=True)
        self.tessWS = self.__tessWB[easygui.choicebox(msg="Which Sheet for Tessetura?",
                                                      title="Horizon Tess/QB Compare",
                                                      choices=self.__tessWB.get_sheet_names())]
        self.qbWS = self.__qbWB[easygui.choicebox(msg="Which Sheet for QB?",
                                                  title="Horizon Tess/QB Compare",
                                                  choices=self.__qbWB.get_sheet_names())]
        self.tessHead = []
        x = 1
        while x <= self.tessWS.max_column:
            cell = self.tessWS.cell(row=1, column=x)
            if cell.value is not None:
                self.tessHead.append([x, cell.value])
                print(cell.value)
            x += 1
        self.qbHead = []
        x = 1
        while x <= self.qbWS.max_column:
            cell = self.qbWS.cell(row=1, column=x)
            if cell.value is not None:
                self.qbHead.append([x, cell.value])
                print(cell.value)
            x += 1
        self.tessCust = ast.literal_eval(easygui.choicebox(msg="Customer ID in Tessetura?",
                                          title="Horizon Tess/QB Compare",
                                          choices=self.tessHead))
        self.qbCust = ast.literal_eval(easygui.choicebox(msg="Customer ID in QB?",
                                          title="Horizon Tess/QB Compare",
                                          choices=self.qbHead))
        self.tessDate = ast.literal_eval(easygui.choicebox(msg="Transaction Date in Tessetura?",
                                          title="Horizon Tess/QB Compare",
                                          choices=self.tessHead))
        self.qbDate = ast.literal_eval(easygui.choicebox(msg="Transaction Date in QB?",
                                          title="Horizon Tess/QB Compare",
                                          choices=self.qbHead))
        self.tessMoney = ast.literal_eval(easygui.choicebox(msg="Money Amount in Tessetura?",
                                          title="Horizon Tess/QB Compare",
                                          choices=self.tessHead))
        self.qbMoney = ast.literal_eval(easygui.choicebox(msg="Money Amount in QB?",
                                          title="Horizon Tess/QB Compare",
                                          choices=self.qbHead))


if __name__ == "__main__":
    q = Horizon()
    print(q.tessHead)
    print(q.qbHead)
    print(q.tessCust[0])
    print(q.qbCust[0])
    print(q.tessDate[0])
    print(q.qbDate[0])
    print(q.tessMoney[0])
    print(q.qbMoney[0])
    print(int("R Owens #8101684"))
